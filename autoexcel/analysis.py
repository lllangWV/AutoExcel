import pandas as pd

import pandas as pd
from openpyxl import load_workbook

def extract_disjoint_tables(file_path, sheet_name=None):
    # Load the workbook using openpyxl
    wb = load_workbook(file_path, data_only=True)
    
    # If no sheet name provided, default to the first sheet
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    
    sheet = wb[sheet_name]
    
    # List to store the tables as dataframes
    tables = []
    
    # Convert sheet to a pandas DataFrame, expanding merged cells
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append([cell if cell is not None else None for cell in row])
    df = pd.DataFrame(data)
    
    # Find the merged cells and apply the values to all merged ranges
    for merged_cells in sheet.merged_cells.ranges:
        merged_value = sheet.cell(merged_cells.min_row, merged_cells.min_col).value
        df.iloc[merged_cells.min_row-1:merged_cells.max_row, merged_cells.min_col-1:merged_cells.max_col] = merged_value

    df_mask = df.notna()


    # Initialize variables
    tables = []
    visited = pd.DataFrame(False, index=df.index, columns=df.columns)
    
    # Function to find contiguous non-NaN regions
    def find_table(start_row, start_col):
        from collections import deque
        q = deque()
        q.append((start_row, start_col))
        table_cells = []
        
        while q:
            row, col = q.popleft()
            if (0 <= row < df.shape[0] and
                0 <= col < df.shape[1] and
                df_mask.iloc[row, col] and
                not visited.iloc[row, col]):
                visited.iloc[row, col] = True
                table_cells.append((row, col))
                
                # Explore neighbors
                q.extend([
                    (row+1, col),
                    (row-1, col),
                    (row, col+1),
                    (row, col-1)
                ])
        return table_cells
    
    # Iterate over the DataFrame to find tables
    for row in range(df.shape[0]):
        for col in range(df.shape[1]):
            if df_mask.iloc[row, col] and not visited.iloc[row, col]:
                # Found a new table
                table_cells = find_table(row, col)
                # Extract the table data
                rows = [r for r, c in table_cells]
                cols = [c for r, c in table_cells]
                table_df = df.iloc[min(rows):max(rows)+1, min(cols):max(cols)+1]
                # Clean up the table (remove all-NaN rows and columns)
                table_df = table_df.dropna(axis=0, how='all').dropna(axis=1, how='all')

                table_df = clean_header(table_df)
                table_df = clean_footer(table_df)
                tables.append(table_df.reset_index(drop=True))
    return tables


def clean_header(df):
    first_5_no_duplicates = df.head(4).drop_duplicates()
    # Combine with the rest of the dataframe
    df_result = pd.concat([first_5_no_duplicates, df.iloc[5:]])
    # Reset the index if needed
    df_result = df_result.reset_index(drop=True)
    return df_result

def clean_footer(df):
    last_5_no_duplicates = df.tail(4).drop_duplicates()
    # Combine with the rest of the dataframe
    df_result = pd.concat([df.iloc[:-5], last_5_no_duplicates])
    # Reset the index if needed
    df_result = df_result.reset_index(drop=True)
    return df_result



# Example usage

def main():
    file_path = 'C:/Users/lllang/Desktop/Current_Projects/Auto_Excel/data/processed/Data Analysis 9-19-2024.xlsx'

    fy_analytics_tables = extract_disjoint_tables(file_path, sheet_name='FY 24-25 Analytics')
    caseload_analysis_tables = extract_disjoint_tables(file_path, sheet_name='Caseload Analysis')
    fy_sharepoint_tables = extract_disjoint_tables(file_path, sheet_name='FY 24 SharePoint')
    active_assignments_tables = extract_disjoint_tables(file_path, sheet_name='Active Assignments')
    # Display each table
    for i, df in enumerate(fy_analytics_tables):
        print(f"Table {i+1}:")
        # I drop duplica
        print(df.head())


if __name__ == '__main__':
    main()