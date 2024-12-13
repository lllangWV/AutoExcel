import argparse
from typing import List
import os
import logging
import traceback
from datetime import datetime
import re
import shutil
import time


import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

import win32com

shutil.rmtree(win32com.__gen_path__)

import win32com.client as win32
win32c = win32.constants

from autoexcel import config

logger = logging.getLogger(__name__)


def fy_analysis(raw_dir, processed_dir, assigned_date_filter=[datetime(2023, 7, 1), None], old_negotiators=None):
    """
    Process FY analysis using the most recent raw data file and template file.
    """
    os.makedirs(processed_dir, exist_ok=True)
    
    # Find most recent raw data and processed files
    raw_xlsx = get_latest_file(raw_dir, r'Raw Data (\d{1,2}[-\.]\d{1,2}[-\.]\d{4})\.xlsx')
   
    logger.info(f"Found latest raw file: {raw_xlsx}")
    
    # Output filename using date from raw file
    raw_date = re.search(r'(\d{1,2}[-\.]\d{1,2}[-\.]\d{4})', os.path.basename(raw_xlsx)).group(1)
    formatted_date = raw_date.replace('.', '-')  # Standardize date format
    output_filename = os.path.join(processed_dir, f'Data Analysis {formatted_date}.xlsx')
    
    logger.info(f"Writing output to: {output_filename}")
    # Remove existing output file if it exists
    if os.path.exists(output_filename):
        logger.info(f"Removing existing output file: {output_filename}")
        os.remove(output_filename)
        
    # Find previous week's data analysis file
    previous_week_xlsx = get_latest_file(processed_dir, r'Data Analysis (\d{1,2}[-\.]\d{1,2}[-\.]\d{4})\.xlsx')
    
    if not raw_xlsx or not previous_week_xlsx:
        raise FileNotFoundError("Could not find required input files")
    
    # Read the data from the Excel file
    df = read_data(raw_xlsx)

    # Process the data
    df_processed, df_active_assignments = preprocess_data(df, assigned_date_filter)
    
    # Write the processed data to Excel
    write_output(df_processed, df_active_assignments, output_filename)

    
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = True  # False
    
    wb = excel.Workbooks.Open(os.path.abspath(output_filename))
    create_fy_analytics_ws(wb, old_negotiators=old_negotiators)
    create_caseload_analysis_ws(wb, date=formatted_date, old_negotiators=old_negotiators)
    
    
    wb.Save()
    wb.Close()
    excel.Quit()

    logger.info(f'Process completed. The processed data has been saved to {output_filename}')
    


def get_latest_file(directory, pattern):
    """
    Find the most recent file in directory matching the pattern.
    
    Args:
        directory (str): Directory path to search
        pattern (str): Regex pattern to match filenames
        
    Returns:
        str: Full path to the most recent file, or None if no files found
    """
    logger.debug(f"Searching for latest file in {directory} with pattern {pattern}")
    
    files = []
    for filename in os.listdir(directory):
        match = re.search(pattern, filename)
        if match:
            date_str = match.group(1)
            # Convert date string to datetime, handling both period and hyphen separators
            try:
                date = datetime.strptime(date_str.replace('.', '-'), '%m-%d-%Y')
                files.append((os.path.join(directory, filename), date))
            except ValueError as e:
                logger.warning(f"Couldn't parse date from filename {filename}: {e}")
    
    if not files:
        logger.warning(f"No matching files found in {directory}")
        return None
        
    # Sort by date and return the most recent file
    latest_file = max(files, key=lambda x: x[1])[0]

    return latest_file

def fy_analysis_from_template(raw_xlsx, template_xlsx, processed_dir='fy_analysis_processed', assigned_date_filter=[datetime(2023, 7, 1), None]):
    os.makedirs(processed_dir, exist_ok=True)
    
    # Read the data from the Excel file
    df = read_data(raw_xlsx)

    # Process the data
    df_processed, df_active_assignments = preprocess_data(df, assigned_date_filter)

    # Output filename
    raw_date = os.path.basename(raw_xlsx).split('.')[0].split()[-1]
    output_filename = os.path.join(processed_dir, f'Data Analysis {raw_date}.xlsx')

    # Write the processed data to Excel
    write_output(df_processed, df_active_assignments, output_filename)

    # Copy worksheets

    fiscal_year, _ = get_current_fiscal_year()
    fy_analytics_ws_name= f'FY {str(fiscal_year-1)[-2:]}-{str(fiscal_year)[-2:]} Analytics'
    copy_excel_worksheet(template_xlsx, output_filename, worksheet_names=[fy_analytics_ws_name, 'Caseload Analysis'])

    logger.info(f'Process completed. The processed data has been saved to {output_filename}')
    


def read_data(raw_xlsx):
    """Reads data from the raw Excel file."""
    df = pd.read_excel(raw_xlsx)
    return df

def preprocess_data(df, assigned_date_filter):
    """Processes the DataFrame according to specified steps."""
    logger.info('Preprocessing data.')
    df = df.copy()

    # Insert sequential numbers
    # df.insert(0, '#', range(1, len(df) + 1))
    df.insert(0, '#', [1] * len(df))

    # Add new columns
    new_columns = ['Time to Assignment', 'Time to Execution', "Today's Date", 'Time Since Assignment', 'Delinquency']
    for col in new_columns:
        df[col] = None

    # Ensure date columns are in datetime format
    date_cols = ['Date Assigned', 'Date Received at OSP', 'FE Date']
    for col in date_cols:
        df[col] = pd.to_datetime(df[col])

    # Filter data based on assigned date
    df = filter_assigned_date(df, start_fiscal_year=assigned_date_filter[0], end_fiscal_year=assigned_date_filter[1])

    # Exclude certain 'Negotiator' values
    df = df[~df['Negotiator'].isin(['COE', 'NCE', 'OGC'])]

    # Calculate time metrics
    df['Time to Assignment'] = df.apply(lambda row: networkdays(row['Date Received at OSP'], row['Date Assigned']), axis=1)
    df['Time to Execution'] = df.apply(lambda row: networkdays(row['Date Assigned'], row['FE Date']) if pd.notna(row['FE Date']) else None, axis=1)
    df["Today's Date"] = pd.to_datetime('today').normalize()
    df['Time Since Assignment'] = df.apply(lambda row: networkdays(row['Date Assigned'], row["Today's Date"]) 
                                         if not row['Status'] in ['Completed', 'Duplicate', 'Withdrawn'] else None, axis=1)

    # Format date columns
    date_cols_formatted = ['Date Assigned', 'Deadline Date', 'Date Received at OSP', 'Date Received at WVU', 'FE Date', "Today's Date"]
    for col in date_cols_formatted:
        if col in df.columns:
            df[col] = df[col].dt.strftime('%m/%d/%Y')

    # Sort and categorize
    df = df.sort_values(by='Time Since Assignment', ascending=False)
    df['Delinquency'] = df['Time Since Assignment'].apply(categorize_delinquency)

    # Create copies for different outputs
    df_original = df.copy()
    df_active_assignments = df_original.copy()
    df_active_assignments = df_active_assignments[~df_active_assignments['Status'].isin(['Completed', 'Duplicate', 'Withdrawn'])]

    logger.info('Preprocessing complete.')
    return df_original, df_active_assignments

def get_current_fiscal_year():
    today = datetime.today()
    if today.month >= 10:
        fiscal_year_start = datetime(today.year, 10, 1)
        fiscal_year = today.year + 1
    else:
        fiscal_year_start = datetime(today.year - 1, 10, 1)
        fiscal_year = today.year
    return fiscal_year, fiscal_year_start

def filter_assigned_date(df, start_fiscal_year, end_fiscal_year):
    logger.debug('Filtering assigned date.')
    df = df[df['Date Assigned'] >= start_fiscal_year]
    if end_fiscal_year:
        df = df[df['Date Assigned'] < end_fiscal_year]
    logger.debug('Filtering assigned date complete.')
    return df

def networkdays(start_date, end_date):
    if pd.isnull(start_date) or pd.isnull(end_date):
        return None
    
    # Convert dates to datetime.date objects in correct format
    start = pd.to_datetime(start_date).date()
    end = pd.to_datetime(end_date).date()

    return np.busday_count(start, end) + 1  # Include end date

def categorize_delinquency(days):
    if pd.isnull(days):
        return None
    elif days >= 90:
        return '> 90 Days'
    elif days >= 60:
        return '> 60 Days'
    elif days >= 30:
        return '> 30 Days'
    else:
        return '< 30 Days'

def write_output(df_original, df_active_assignments, output_filename):
    """Writes the processed data to an Excel file with formatting."""
    
    # logger.debug('Writing output to Excel.')
    logger.info(f"Writing output to {output_filename}")
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        fiscal_year, _ = get_current_fiscal_year()
        fy_sheet_name = f'FY {str(fiscal_year-1)[-2:]}-{str(fiscal_year)[-2:]} SharePoint'

        # Write data to worksheets
        # Convert date columns to datetime before writing
        date_columns = ['Date Assigned', 'FE Date']
        
        df_original_copy = df_original.copy()
        df_active_copy = df_active_assignments.copy()
        
        for col in date_columns:
            if col in df_original_copy.columns:
                df_original_copy[col] = pd.to_datetime(df_original_copy[col])
            if col in df_active_copy.columns:
                df_active_copy[col] = pd.to_datetime(df_active_copy[col])
                
        df_original_copy.to_excel(writer, index=False, sheet_name=fy_sheet_name)
        df_active_copy.to_excel(writer, index=False, sheet_name="Active Assignments")
        
        # Apply formatting
        for sheet_name, df in zip([fy_sheet_name, "Active Assignments"], [df_original, df_active_assignments]):
            worksheet = writer.sheets[sheet_name]
            format_worksheet(worksheet, df)
            
    logger.debug("Writing output to Excel complete.")
    
def format_worksheet(worksheet, df):
    """Applies formatting to the Excel worksheet."""
    logger.debug('Formatting worksheet.')
    # Set default row height
    worksheet.sheet_format.defaultRowHeight = 15

    # Apply auto-filter
    max_column = get_column_letter(worksheet.max_column)
    worksheet.auto_filter.ref = f"A1:{max_column}1"

    # Format header
    title_fill = PatternFill(start_color="C0E6F4", end_color="C0E6F4", fill_type="solid")
    non_bold_font = Font(bold=False)
    
    # Apply wrap text to all cells while keeping row height at 15
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrapText=True)
            cell.parent.row_dimensions[cell.row].height = 15
            
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = title_fill
        cell.font = non_bold_font

    # Apply conditional formatting for 'Delinquency'
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    orange_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    if 'Delinquency' in df.columns:
        delinquency_col_idx = df.columns.get_loc('Delinquency') + 1
        delinquency_col_letter = get_column_letter(delinquency_col_idx)

        for idx, cell in enumerate(worksheet[delinquency_col_letter][1:], start=2):
            delinquency_value = cell.value
            if delinquency_value == '> 90 Days':
                cell.fill = red_fill
            elif delinquency_value == '> 60 Days':
                cell.fill = orange_fill
            elif delinquency_value == '> 30 Days':
                cell.fill = yellow_fill
    logger.debug('Formatting worksheet complete.')

def copy_excel_worksheet(source_excel_path, target_excel_path, worksheet_names: List[str]):
    """Copies worksheets from the source Excel file to the target Excel file."""
    logger.debug('Copying worksheets.')
    from win32com.client import Dispatch
    # print(dir(win32c))
    xl = Dispatch("Excel.Application")
    xl.Visible = False

    fiscal_year, _ = get_current_fiscal_year()
    
    soruce_wb = xl.Workbooks.Open(Filename=os.path.abspath(source_excel_path))
    target_wb = xl.Workbooks.Open(Filename=os.path.abspath(target_excel_path))
    logger.debug(f"Attempting to copy worksheets from {source_excel_path} to {target_excel_path}")
    try:
        for worksheet_name in worksheet_names:
            logger.debug(f"Attempting to copy worksheet {worksheet_name}")
            source_ws = soruce_wb.Worksheets(worksheet_name)
            source_ws.Copy(Before=target_wb.Worksheets(1))
            logger.debug(f"Successfully copied worksheet {worksheet_name}")
            
    except Exception as e:
        logger.exception(e)
    finally:
        soruce_wb.Close(SaveChanges=False)
        target_wb.Close(SaveChanges=True)
        xl.Quit()

    logger.debug('Copying worksheets complete.')
    
def pivot_table(wb: object, 
                ws1: object, 
                pt_ws: object, 
                ws_name: str, 
                pt_name: str, 
                pt_rows: list, 
                pt_cols: list, 
                pt_filters: list, 
                pt_fields: list, 
                start_row: int = 1, 
                start_col: int = 1,
                apply_grouping: bool = False,

                pt_visible_rows: list = None,
                footer_text: str = None,
                visible_items: dict = None,
                show_detail_items: dict = None,
                sorting_items: dict = None,
                header_size: tuple = (2,0)):
    """
    wb = workbook1 reference
    ws1 = worksheet1
    pt_ws = pivot table worksheet number
    ws_name = pivot table worksheet name
    pt_name = name given to pivot table
    pt_rows, pt_cols, pt_filters, pt_fields: values selected for filling the pivot tables
    start_row = starting row number for pivot table (default 1)
    start_col = starting column number for pivot table (default 1)
    apply_grouping = boolean to apply grouping to date fields (default False)
    pt_visible_rows = list of rows to make visible in pivot table (default [])
    footer_text = text to add to the footer of the pivot table (default None)
    visible_items = dictionary of items to make visible in pivot table (default None)
    show_detail_items = dictionary of items to show detail for in pivot table (default None)
    """
    header_row_size = header_size[0]
    header_col_size = header_size[1]
    
    # pivot table location - add 1 to account for title
    pt_loc_row = start_row + len(pt_filters) + header_row_size
    
    # grab the pivot table source data
    pc = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=ws1.UsedRange)
    
    # create the pivot table object
    pc.CreatePivotTable(TableDestination=f'{ws_name}!R{pt_loc_row}C{start_col}', TableName=pt_name)

    # selecte the pivot table work sheet and location to create the pivot table
    pt_ws.Select()
    pt_ws.Cells(pt_loc_row, start_col).Select()
    
    
    # Grouping date fields
    for field_list, field_r in ((pt_filters, win32c.xlPageField), (pt_rows, win32c.xlRowField), (pt_cols, win32c.xlColumnField)):
        for i, value in enumerate(field_list):
            pf = pt_ws.PivotTables(pt_name).PivotFields(value)
            pf.Orientation = field_r
            pf.Position = i + 1
        
            # if value in grouping_list:
            #     logger.debug(f"Applying grouping to field: {value}")
            #     try:
            #         pf.AutoGroup()
            #     except Exception as e:
            #         logger.debug(e)
            #         logger.debug(f"Could not group field: {value}")
            if apply_grouping:
                if any(date_key in value.lower() for date_key in ['date', 'months', 'years']):
                    
                    try:
                    # Group by months, quarters, and years
                        pf.AutoGroup()  # Auto group the dates
                    except Exception as e:
                        logger.debug(e)
                        logger.debug(f"Could not group field: {value}")
    # List all field names
    logger.debug("All available pivot table fields:")
    for field in pt_ws.PivotTables(pt_name).PivotFields():
        logger.debug(f"Field name: {field.Name}")
        
    # Visible Items. This will select and deselect items in the pivot table
    if visible_items:       
        for field_name, field_items in visible_items.items():
            for item in field_items:
                item_name = item[0] # String
                item_show_detail = item[1] # boolean
                logger.debug(f"Field Name: {field_name}, Item Name: {item_name}, Show Detail: {item_show_detail}")
                try:
                    pt_ws.PivotTables(pt_name).PivotFields(field_name).PivotItems(item_name).Visible = item_show_detail
                except Exception as e:
                    logger.debug(e)
                    logger.debug(f"Could not set visible item: {item_name}")
    
    # Hide all fields
    for field in pt_ws.PivotTables(pt_name).VisibleFields:
        name = field.Name
        logger.debug(f"Hiding field: {name}")
        try:
            pf = pt_ws.PivotTables(pt_name).PivotFields(name)
            pf.Orientation = win32c.xlHidden 
        except Exception as e:
            logger.debug(e)
            logger.debug(f"Could not hide field: {name}")
    
    # Set the visible rows
    if pt_visible_rows is None:
        pt_visible_rows = pt_rows
        
    logger.debug(f"Setting visible rows: {pt_visible_rows}")
    for x in pt_visible_rows:
        logger.debug(f"Setting visible row: {x}")
        try:
            pf = pt_ws.PivotTables(pt_name).PivotFields(x)
            pf.Orientation = win32c.xlRowField 
        except Exception as e:
            logger.debug(e)
            logger.debug(f"Could not set visible row: {x}")
    
    
    # Show Detail Items. This will show or hide the detail rows in the pivot table
    if show_detail_items:
        for field_name, field_items in show_detail_items.items():
            for item in field_items:
                item_name = item[0] # String
                item_show_detail = item[1] # boolean
                logger.debug(f"Field Name: {field_name}, Item Name: {item_name}, Show Detail: {item_show_detail}")
                try:
                    pt_ws.PivotTables(pt_name).PivotFields(field_name).PivotItems(item_name).ShowDetail = item_show_detail
                except Exception as e:
                    logger.debug(e)
                    logger.debug(f"Could not set show detail item: {item_name}")
                    

    # Sets the Values of the pivot table
    for field in pt_fields:
        try:
            pt_ws.PivotTables(pt_name).AddDataField(pt_ws.PivotTables(pt_name).PivotFields(field[0]), field[1], field[2]).NumberFormat = field[3]
        except Exception as e:
            logger.debug(e)
            logger.debug(f"Could not add data field: {field[0]}")
    
    pt_ws.PivotTables(pt_name).ShowValuesRow = False
    pt_ws.PivotTables(pt_name).ColumnGrand = True
    
    # Sorting items after applying Values
    if sorting_items:
        for field_name, sorting_item in sorting_items.items():
            sorting_order = sorting_item[1]
            sorting_field = sorting_item[0]
            pt_ws.PivotTables(pt_name).PivotFields(field_name).AutoSort(sorting_order, sorting_field)
        
    # Get the dimensions of the pivot table
    width = pt_ws.PivotTables(pt_name).TableRange1.Columns.Count
    height = pt_ws.PivotTables(pt_name).TableRange1.Rows.Count
    
    
    logger.debug(f"Pivot table dim: {width} x {height}")
    logger.debug(f"Start Row: {start_row}, Start Col: {start_col}")
    logger.debug(start_col + width)
    
    logger.debug("Formating Table")
    
    # Add Header to the pivot table
    header_range = pt_ws.Range(pt_ws.Cells(start_row, start_col), pt_ws.Cells(start_row + 1, start_col + width - 1))
    header_range.Merge()
    header_range.Value = pt_name
    header_range.Font.Bold = True
    header_range.Font.Size = 12
    header_range.Font.Name = "Aptos Narrow"
    header_range.Font.ColorIndex = win32c.xlThemeColorAccent6  # Green Accent 6
    header_range.HorizontalAlignment = win32c.xlCenter  # Center align
    header_range.VerticalAlignment = win32c.xlCenter  # Middle align
    header_range.WrapText = True  # Wrap text
    
    # Add black border to header
    header_range.Borders.LineStyle = win32c.xlContinuous
    header_range.Borders.Color = 0x000000  # Black
    header_range.Borders.Weight = win32c.xlThin
    height+=2
    # Merge cells across the table width
    
    # Add footer below pivot table
    if footer_text:
        pt = pt_ws.PivotTables(pt_name)
        lastRow = pt.TableRange2.Row + pt.TableRange2.Rows.Count
        
        footer_range = pt_ws.Range(pt_ws.Cells(lastRow, start_col), pt_ws.Cells(lastRow , start_col + width - 1))
        footer_range.Merge()
        footer_range.Value = footer_text
        footer_range.Font.Size = 11
        footer_range.Font.Name = "Aptos Narrow"
        footer_range.HorizontalAlignment = win32c.xlLeft
        footer_range.Borders.LineStyle = win32c.xlContinuous
        footer_range.Borders.Color = 0x000000  # Black
        footer_range.Borders.Weight = win32c.xlThin
        height+=1
    
    # Add black borders to all cells in pivot table
    pt_range = pt_ws.PivotTables(pt_name).TableRange2
    pt_range.Borders.LineStyle = win32c.xlContinuous
    pt_range.Borders.Color = 0x000000  # Black
    pt_range.Borders.Weight = win32c.xlThin
    
    # pt_ws.PivotTables(pt_name)
    
    return (width, height)


def create_caseload_analysis_ws(wb, date=None, old_negotiators = None):
    ws_active = wb.Sheets('Active Assignments')
    if date is None:
        date = datetime.today().strftime('%m/%d/%Y')
        
    ws_caseload_name = "Caseload_Analysis"
    wb.Sheets.Add().Name = ws_caseload_name
    ws_caseload = wb.Sheets(ws_caseload_name)


    current_row=1
    current_col=1
    seperation_of_pivot_tables = 2

    pt_name = f'Caseload Analysis {date}'
    pt_rows = ['Negotiator']
    pt_cols = []
    pt_filters = []

    show_detail_items = None
    visible_items={}
    if old_negotiators:
        visible_items['Negotiator'] = [(negotiator, False) for negotiator in old_negotiators]
        visible_items['FE Date'] = [('(blank)', False)]
        visible_items['High Priority'] = [('(blank)', False)]

    pt_fields = [['#', 'Open Cases in SharePoint', win32c.xlSum, '0']]

    logger.info(f"Creating pivot table: {pt_name}")
    pivot_table(wb, ws_active, ws_caseload, ws_caseload_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields, 
                start_row=current_row, start_col=current_col, 
                show_detail_items=show_detail_items, visible_items=visible_items)

    ws_caseload.Columns.AutoFit()
    

def create_fy_analytics_ws(wb, old_negotiators = None):
    ws_active = wb.Sheets('Active Assignments')
    ws_sharepoint = wb.Sheets('FY 24-25 SharePoint')
    # ws2_name = "FY_24-25_Analytics"
    ws2_name = "FY_24_25_Analytics"
    wb.Sheets.Add().Name = ws2_name
    ws2 = wb.Sheets(ws2_name)
    
    visible_items={}
    if old_negotiators:
        visible_items['Negotiator'] = [(negotiator, False) for negotiator in old_negotiators]
        visible_items['FE Date'] = [('(blank)', False)]
        visible_items['High Priority'] = [('(blank)', False)]
        
    logger.info(f"Unselecting old negotiators: {old_negotiators}")
        

    ####################################################################################################################################################
    current_row=1
    current_col=1
    seperation_of_pivot_tables = 2

    pt_name = 'Active Caseload by Negotiator, Status, Agreement Type, and Delinquency'
    pt_rows = ['Negotiator', 'Status', 'Agreement Type', 'Delinquency']
    grouping_list = ['FE Date', 'Date Assigned']
    pt_cols = []
    pt_filters = []

    show_detail_items = {'Months (Date Assigned)': [('Jan',False), ('Feb',False), ('Mar',False), ('Apr',False), 
                                        ('May',False), ('Jun',False), ('Jul',False), ('Aug',False), 
                                        ('Sep',False), ('Oct',False), ('Nov',False), ('Dec',False)]}
                        # 'Years (Date Assigned)': [('2023',False)]}


    pt_fields = [['#', 'Sum of #', win32c.xlSum, '0']]

    logger.info(f"Creating pivot table: {pt_name}")
    pivot_table(wb, ws_active, ws2, ws2_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields, 
                start_row=current_row, start_col=current_col, apply_grouping=True,
                show_detail_items=show_detail_items, visible_items=visible_items)

    ####################################################################################################################################################

    current_row=1
    current_col=4
    seperation_of_pivot_tables = 2

    pt_name = 'Executed Agreements by Negotiator'
    pt_rows = ['Negotiator', 'FE Date', 'Date Assigned']
    # grouping_list = ['FE Date', 'Date Assigned']
    pt_cols = []
    pt_filters = []
    pt_fields = [['#', 'Sum of #', win32c.xlSum, '0']]
    pt_visible_rows = ['Negotiator', 'Years (FE Date)',  'Months (FE Date)', 'FE Date']

    show_detail_items = {'Months (FE Date)': [('Jan',False), ('Feb',False), ('Mar',False), ('Apr',False), 
                                        ('May',False), ('Jun',False), ('Jul',False), ('Aug',False), 
                                        ('Sep',False), ('Oct',False), ('Nov',False), ('Dec',False)]}
    
    # Get unique FE dates from SharePoint worksheet
    fe_dates = set()
    fe_date_col = None
    for col in range(1, ws_sharepoint.UsedRange.Columns.Count + 1):
        if ws_sharepoint.Cells(1, col).Value == "FE Date":
            fe_date_col = col
            break
    
    if fe_date_col:
        for row in range(2, ws_sharepoint.UsedRange.Rows.Count + 1):
            date_val = ws_sharepoint.Cells(row, fe_date_col).Value
            if date_val not in fe_dates and date_val:
                fe_dates.add(str(date_val.year))
    

    max_year = 0
    for year in fe_dates:
        max_year = max(max_year, int(year))
    show_detail_items['Years (FE Date)'] = [(str(year), year == str(max_year)) for year in fe_dates]
    
    logger.info(f"Creating pivot table: {pt_name}")
    pivot_table(wb, ws_sharepoint, ws2, ws2_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields, 
                start_row=current_row, start_col=current_col, apply_grouping=True, pt_visible_rows=pt_visible_rows,
                show_detail_items=show_detail_items, visible_items=visible_items)


    ####################################################################################################################################################

    current_row=1
    current_col=7
    seperation_of_pivot_tables = 2

    pt_name = 'New Assignments by Negotiator'
    pt_rows = ['Negotiator', 'FE Date', 'Date Assigned']
    pt_cols = []
    pt_filters = []
    pt_fields = [['#', 'Sum of #', win32c.xlSum, '0']]
    pt_visible_rows = ['Negotiator', 'Years (Date Assigned)',  'Months (Date Assigned)', 'Date Assigned']

    show_detail_items = {'Months (Date Assigned)': [('Jan',False), ('Feb',False), ('Mar',False), ('Apr',False), 
                                        ('May',False), ('Jun',False), ('Jul',False), ('Aug',False), 
                                        ('Sep',False), ('Oct',False), ('Nov',False), ('Dec',False)]}
    
    date_assigned_dates = set()
    date_assigned_col = None
    for col in range(1, ws_sharepoint.UsedRange.Columns.Count + 1):
        if ws_sharepoint.Cells(1, col).Value == "Date Assigned":
            date_assigned_col = col
            break
    
    if date_assigned_col:
        for row in range(2, ws_sharepoint.UsedRange.Rows.Count + 1):
            date_val = ws_sharepoint.Cells(row, date_assigned_col).Value
            if date_val not in date_assigned_dates and date_val:
                date_assigned_dates.add(str(date_val.year))

    max_year = 0
    for year in date_assigned_dates:
        max_year = max(max_year, int(year))
    show_detail_items['Years (Date Assigned)'] = [(str(year), year == str(max_year)) for year in date_assigned_dates]
 
    logger.info(f"Creating pivot table: {pt_name}")
    pivot_table(wb, ws_sharepoint, ws2, ws2_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields, 
                start_row=current_row, start_col=current_col, apply_grouping=True, pt_visible_rows=pt_visible_rows, 
                visible_items=visible_items, show_detail_items=show_detail_items)

    ####################################################################################################################################################

    current_row=1
    current_col=10
    seperation_of_pivot_tables = 2

    pt_name = 'Delinquency by Negotiator'
    pt_rows = ['Negotiator', 'Delinquency']
    pt_cols = []
    pt_filters = []
    pt_fields = [['#', 'Sum of #', win32c.xlSum, '0']]
    visible_items = {'Negotiator': [('Abigail Gallagher', False), ('Eric Divito', False), 
                                    ('Eric Winaught', False), ('Jillian Corbett', False), ('Huron – New', False)],
                    'FE Date': [('(blank)', False)],
                    'High Priority': [('(blank)', False)]}

    prev_width, prev_height = pivot_table(wb, ws_active, ws2, ws2_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields, 
                                        start_row=current_row, start_col=current_col, 
                                        visible_items=visible_items)
    current_row += prev_height + seperation_of_pivot_tables

    pt_name = 'Total Assignments By Delinquency'
    pt_rows = [ 'Delinquency']
    pt_cols = []
    pt_filters = []
    pt_fields = [['#', 'Sum of #', win32c.xlSum, '0']]

    prev_width, prev_height = pivot_table(wb, ws_active, ws2, ws2_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields, 
                                        start_row= current_row, start_col=current_col, visible_items=visible_items)
    current_row += prev_height + seperation_of_pivot_tables

    pt_name = 'Average Time to Assignment by Month'
    pt_rows = [ 'Delinquency', 'FE Date', 'Date Assigned']
    pt_cols = []
    pt_filters = []
    pt_fields = [['Time to Assignment', 'Average of TTA*', win32c.xlAverage, '0']]
    pt_visible_rows = ['Years (Date Assigned)',   'Months (Date Assigned)', 'Date Assigned']
    footer_text = "*Data is skewed due to reassignments"

    show_detail_items = {'Months (Date Assigned)': [('Jan',False), ('Feb',False), ('Mar',False), ('Apr',False), 
                                        ('May',False), ('Jun',False), ('Jul',False), ('Aug',False), 
                                        ('Sep',False), ('Oct',False), ('Nov',False), ('Dec',False)]}

    logger.info(f"Creating pivot table: {pt_name}")
    pivot_table(wb, ws_sharepoint, ws2, ws2_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields, 
                start_row= current_row, start_col=current_col, apply_grouping=True, pt_visible_rows=pt_visible_rows, footer_text=footer_text,
                visible_items=visible_items, show_detail_items=show_detail_items)



    ####################################################################################################################################################

    current_row=1
    current_col=13
    seperation_of_pivot_tables = 2
    pt_name = 'Time to Assignment and Time to Completion Analysis by Negotiator'
    pt_rows = ['Negotiator', 'FE Date', 'Date Assigned', 'Time to Assignment', 'Time to Execution']
    pt_cols = []
    pt_filters = []
    pt_fields = [['Time to Assignment', 'Average of TTA*', win32c.xlAverage, '0'],
                ['Time to Execution', 'Average of TTE*', win32c.xlAverage, '0']]
    pt_visible_rows = ['Negotiator']

    show_detail_items = {'Negotiator': [('David WomoChil',False), ('Jocelyn Phares',False), 
                                        ('Julie Bonasso',False), ('Justin Miller',False), 
                                        ('Laken Dillon',False), ('Matthew Nesmith',False), 
                                        ('Rachel Hanisch',False), ('Stephanie Harrod',False), 
                                        ('Waynell Henson',False)]}

    footer_text = "* TTA = Time to Assignment, TTE = Time to Execution"
    logger.info(f"Creating pivot table: {pt_name}")
    prev_width, prev_height = pivot_table(wb, ws_sharepoint, ws2, ws2_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields, 
                start_row=current_row, start_col=current_col, apply_grouping=True, pt_visible_rows=pt_visible_rows, footer_text=footer_text,
                show_detail_items=show_detail_items, visible_items=visible_items)
    current_row += prev_height + seperation_of_pivot_tables




    pt_name = 'High Priority Assignment Analysis by Negotiator'
    pt_rows = ['Negotiator', 'FE Date', 'Date Assigned', 'Time to Assignment', 'Time to Execution', 'High Priority']
    pt_cols = []
    pt_filters = []
    pt_fields = [['#', 'Sum of #', win32c.xlSum, '0'],
                ['Time to Assignment', 'Average of TTA*', win32c.xlAverage, '0'],
                ['Time to Execution', 'Average of TTE*', win32c.xlAverage, '0']]
    pt_visible_rows = ['Negotiator', 'High Priority']
    footer_text = "* TTA = Time to Assignment, TTE = Time to Execution"

    show_detail_items = {'Negotiator': [('David WomoChil',False), ('Jocelyn Phares',False), 
                                        ('Julie Bonasso',False), ('Justin Miller',False), 
                                        ('Laken Dillon',False), ('Matthew Nesmith',False), 
                                        ('Rachel Hanisch',False), ('Stephanie Harrod',False), 
                                        ('Waynell Henson',False)]}

    # show_detail_items = 
    logger.info(f"Creating pivot table: {pt_name}")
    prev_width, prev_height = pivot_table(wb, ws_sharepoint, ws2, ws2_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields, 
                                        start_row=current_row, start_col=current_col, apply_grouping=True, pt_visible_rows=pt_visible_rows, footer_text=footer_text, 
                                        show_detail_items=show_detail_items, visible_items=visible_items)
    current_row += prev_height + seperation_of_pivot_tables

    pt_name = 'Average Time to Execution by Agreement Type'
    pt_rows = ['Negotiator', 'FE Date', 'Date Assigned', 'Time to Execution', 'Agreement Type']
    pt_cols = []
    pt_filters = []
    pt_fields = [['Time to Execution', 'Average of TTE*', win32c.xlAverage, '0']]
    pt_visible_rows = [ 'Agreement Type']
    footer_text = "* TTE = Time to Execution"

    sorting_items = {'Agreement Type': ('Average of TTE*', win32c.xlAscending)}

    logger.info(f"Creating pivot table: {pt_name}")
    prev_width, prev_height = pivot_table(wb, ws_sharepoint, ws2, ws2_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields, 
                                        start_row=current_row, start_col=current_col, apply_grouping=True, pt_visible_rows=pt_visible_rows, footer_text=footer_text,
                                        show_detail_items=show_detail_items, visible_items=visible_items, sorting_items=sorting_items)
    current_row += prev_height

    ws2.Columns.AutoFit()
    
    
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Run AutoExcel processing script.")
    
    data_dir=os.path.join(config.data_dir, 'fy_analysis')
    raw_dir = os.path.join(data_dir, 'raw')
    template_dir = os.path.join(data_dir, 'templates')
    
    
    # Add arguments
    parser.add_argument('--raw_dir', type=str, default=raw_dir, help="Relative path to the raw data directory.")
    parser.add_argument('--template_dir', type=str, default=template_dir, help="Relative path to the template directory.")
    parser.add_argument('--processed_dir', type=str, default='.', help="The output directory for the processed Excel file.")
    parser.add_argument('--data_dir', type=str, default=data_dir, help="Path to the data directory.")
    parser.add_argument('--start_date', type=str, default=None, help="Start date for filtering in YYYY-MM-DD format.")
    parser.add_argument('--end_date', type=str, default=None, help="End date for filtering in YYYY-MM-DD format.")

    # Parse arguments
    args = parser.parse_args()
    
    if args.start_date:
        print(*[int(x) for x in args.start_date.split('-')])
        start_date = datetime(*[int(x) for x in args.start_date.split('-')])
    
    end_date=None
    if args.end_date:
        end_date = datetime(*[int(x) for x in args.end_date.split('-')])
        
    assigned_date_filter=[start_date, end_date]

    fy_analysis(raw_dir, template_dir, processed_dir=args.processed_dir, assigned_date_filter=assigned_date_filter)