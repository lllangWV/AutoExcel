from typing import List

import os
import shutil
import traceback
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


# Parameters
raw_xlsx=os.path.join('data','raw','Raw Data 9-19-2024.xlsx')
processed_xlsx=os.path.join('data','templates','template_processed_workbook.xlsx')
test_dir= os.path.join('data','pandas_test')
data_dir='data'
processed_dir=os.path.join('data','processed')
assigned_date_filter=[datetime(2023, 7, 1),None]


shutil.rmtree(test_dir, ignore_errors=True)
os.makedirs(test_dir)


###################################################################
# Script starts here
###################################################################

# Step 1: Read the data from the Excel file
df = pd.read_excel(raw_xlsx)

raw_date = os.path.basename(raw_xlsx).split('.')[0].split()[-1]


# Step 2: Insert a new first column "#" with sequential numbers starting from 1
df.insert(0, '#', range(1, len(df) + 1))

# Step 3: Add new columns
df['Time to Assignment'] = None
df['Time to Execution'] = None
df["Today's Date"] = None
df['Time Since Assignment'] = None
df['Delinquency'] = None

# Ensure date columns are in datetime format
df['Date Assigned'] = pd.to_datetime(df['Date Assigned'])
df['Date Received at OSP'] = pd.to_datetime(df['Date Received at OSP'])
df['FE Date'] = pd.to_datetime(df['FE Date'])


def get_current_fiscal_year():
    today = datetime.today()
    if today.month >= 10:
        fiscal_year_start = datetime(today.year, 10, 1)
        fiscal_year = today.year + 1
    else:
        fiscal_year_start = datetime(today.year - 1, 10, 1)
        fiscal_year = today.year
    return fiscal_year, fiscal_year_start


# Step 4: Filter out previous fiscal year data based on "Date Assigned"
def filter_assigned_date(df, start_fiscal_year=datetime(2023, 7, 1), end_fiscal_year=None):
    df = df[df['Date Assigned'] >= start_fiscal_year]
    if end_fiscal_year:
        df = df[df['Date Assigned'] < end_fiscal_year]
    return df

df = filter_assigned_date(df, start_fiscal_year=assigned_date_filter[0], end_fiscal_year=assigned_date_filter[1])

# Step 5: Keep only rows where 'Negotiator' is 'COE', 'NCE', or 'OGC'
df = df[~df['Negotiator'].isin(['COE', 'NCE', 'OGC'])]



# Step 7: Define function to calculate business days
def networkdays(start_date, end_date):
    if pd.isnull(start_date) or pd.isnull(end_date):
        return None
    return np.busday_count(start_date.date(), end_date.date()) + 1  # Include end date

# Step 8: Calculate 'Time to Assignment'
df['Time to Assignment'] = df.apply(lambda row: networkdays(row['Date Received at OSP'], row['Date Assigned']), axis=1)

# Step 9: Calculate 'Time to Execution'
df['Time to Execution'] = df.apply(lambda row: networkdays(row['Date Assigned'], row['FE Date']), axis=1)

# Step 10: Set "Today's Date" to current date
df["Today's Date"] = pd.to_datetime('today').normalize()

# Step 11: Calculate 'Time Since Assignment'
df['Time Since Assignment'] = df.apply(lambda row: networkdays(row['Date Assigned'], row["Today's Date"]), axis=1)


# Format the date columns in {month}/{day}/{year} format
df['Date Assigned'] = df['Date Assigned'].dt.strftime('%m/%d/%Y')
df['Deadline Date'] = df['Deadline Date'].dt.strftime('%m/%d/%Y')
df['Date Received at OSP'] = df['Date Received at OSP'].dt.strftime('%m/%d/%Y')
df['Date Received at WVU'] = df['Date Received at WVU'].dt.strftime('%m/%d/%Y')
df['FE Date'] = df['FE Date'].dt.strftime('%m/%d/%Y')
df["Today's Date"] = df["Today's Date"].dt.strftime('%m/%d/%Y')

# Step 12: Sort by 'Time Since Assignment' descending
df = df.sort_values(by='Time Since Assignment', ascending=False)

# Step 13: Categorize 'Delinquency'
def categorize_delinquency(days):
    if pd.isnull(days):
        return None
    elif days >= 90:
        return '> 90 Days'
    elif days >= 60:
        return '> 60 Days'
    elif days >= 30:
        return '> 30 Days'
    else:
        return '< 30 Days'

df['Delinquency'] = df['Time Since Assignment'].apply(categorize_delinquency)

df_original = df.copy()
df_active_assignments = df_original.copy()

# Step 35: In 'Status', deselect 'Completed', 'Duplicate', and 'Withdrawn'
df_active_assignments = df_active_assignments[~df_active_assignments['Status'].isin(['Completed', 'Duplicate', 'Withdrawn'])]



# This only copies values, no formatting
def copy_worksheet(source_sheet, target_sheet):
    # Cpy cell values from source to target sheet
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)
    return None


output_filename=os.path.join(processed_dir,f'Data Analysis {raw_date}.xlsx')
with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    fiscal_year, fiscal_year_start = get_current_fiscal_year()
    
    fy_sheet_name = f'FY {str(fiscal_year)[-2:]} SharePoint'

    # Write the data to worksheets
    df_original.to_excel(writer, index=False, sheet_name=fy_sheet_name)
    df_active_assignments.to_excel(writer, index=False, sheet_name="Active Assignments")

    workbook = writer.book

    # Formatting for the worksheets
    for sheet_name in ["Active Assignments",fy_sheet_name]:
        worksheet = writer.sheets[sheet_name]
        # Set default row height to 15
        worksheet.sheet_format.defaultRowHeight = 15

        # Get the range for the filter (from A1 to the last column in the first row)
        max_column = get_column_letter(worksheet.max_column)
        worksheet.auto_filter.ref = f"A1:{max_column}1"  # Apply filter to the first row

        title_fill = PatternFill(start_color="C0E6F4", end_color="C0E6F4", fill_type="solid")

        # Define a non-bold font
        non_bold_font = Font(bold=False)

        # Apply the fill to all the header cells (1st row)
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)  # Access each cell in the first row
            cell.fill = title_fill  # Apply the fill color
            cell.font = non_bold_font  # Set the font to non-bold

        # Define fill colors for 'Delinquency'
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        orange_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

        # Get the column index of 'Delinquency'
        delinquency_col_idx = df.columns.get_loc('Delinquency') + 1  # Excel columns are 1-indexed
        delinquency_col_letter = get_column_letter(delinquency_col_idx)

        # Apply fill colors based on 'Delinquency' values
        for idx, cell in enumerate(worksheet[delinquency_col_letter][1:], start=2):  # Skip header row
            delinquency_value = cell.value
            if delinquency_value == '> 90 Days':
                cell.fill = red_fill
            elif delinquency_value == '> 60 Days':
                cell.fill = orange_fill
            elif delinquency_value == '> 30 Days':
                cell.fill = yellow_fill
            else:
                pass  # No color for '< 30 Days'

print('Process completed. The processed data has been saved to', output_filename)

def copy_excel_worksheet(source_excel_path, target_excel_path, worksheet_names: List[str]):
    from win32com.client import Dispatch
    xl = Dispatch("Excel.Application")
    xl.Visible = False  # You can remove this line if you don't want the Excel application to be visible

    fiscal_year, fiscal_year_start = get_current_fiscal_year()
    print(os.path.exists(source_excel_path))
    # pywin32 require an absolute path
    wb1 = xl.Workbooks.Open(Filename=os.path.abspath(source_excel_path))
    wb2 = xl.Workbooks.Open(Filename=os.path.abspath(target_excel_path))
    try:
        for worksheet_name in worksheet_names:
            ws1 = wb1.Worksheets(worksheet_name)

            # Copying the the worksheet from the source workbook to the target workbook as the first worksheet in the target workbook
            ws1.Copy(Before=wb2.Worksheets(1))

            if 'FY' in worksheet_name:
                # Formatting for the worksheets
                ws2 = wb2.Worksheets(worksheet_name)
                ws2.Name = f'FY {str(fiscal_year)[-2:]}-{str(fiscal_year+1)[-2:]} Analytics'
                
    except Exception as e:
        print(e)
        traceback.print_exc()

    

    wb1.Close(SaveChanges=False)
    wb2.Close(SaveChanges=True)
    xl.Quit()

copy_excel_worksheet(processed_xlsx, output_filename, worksheet_names=[ 'FY 24-25 Analytics', 'Caseload Analysis'])